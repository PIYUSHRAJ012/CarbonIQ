from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from carbon.services.emission_import.base import ImportedEmissionFactor


class CEASourceAdapter:
    """
    Adapter for the Central Electricity Authority (CEA)
    CO2 Baseline Database Excel workbook.
    """

    SOURCE_NAME = (
        "Central Electricity Authority (India) - "
        "CO2 Baseline Database"
    )

    RESULTS_SHEET = "Results"

    VERSION_CELL = "F4"
    PUBLICATION_DATE_CELL = "F5"

    FISCAL_YEAR_HEADER_ROW = 12
    GRID_FACTOR_ROW = 14

    @classmethod
    def parse_workbook(
        cls,
        workbook_path: str | Path,
    ) -> ImportedEmissionFactor:
        """
        Parse the official CEA workbook and return the latest
        available weighted-average grid emission factor.

        The CEA workbook reports the factor in tCO2/MWh.
        Numerically, this is equivalent to kgCO2/kWh.

        Raises:
            ValueError: If the workbook structure or data is invalid.
        """

        workbook_path = Path(workbook_path)

        if not workbook_path.exists():
            raise ValueError(
                f"CEA workbook not found: {workbook_path}"
            )

        workbook = load_workbook(
            workbook_path,
            data_only=True,
            read_only=True,
        )

        try:
            if cls.RESULTS_SHEET not in workbook.sheetnames:
                raise ValueError(
                    f"Required sheet '{cls.RESULTS_SHEET}' "
                    "was not found in the CEA workbook."
                )

            sheet = workbook[cls.RESULTS_SHEET]

            version = sheet[cls.VERSION_CELL].value
            publication_date = sheet[cls.PUBLICATION_DATE_CELL].value

            if not version:
                raise ValueError(
                    "CEA workbook version could not be determined."
                )

            if not isinstance(publication_date, date):
                raise ValueError(
                    "CEA publication date could not be determined."
                )

            # Find the latest fiscal-year column in row 12.
            latest_column = None
            latest_fiscal_year = None

            for column in range(1, sheet.max_column + 1):
                value = sheet.cell(
                    row=cls.FISCAL_YEAR_HEADER_ROW,
                    column=column,
                ).value

                if not isinstance(value, str):
                    continue

                if "-" not in value:
                    continue

                try:
                    start_year = int(value.split("-")[0])
                except (ValueError, IndexError):
                    continue

                if (
                    latest_fiscal_year is None
                    or start_year > latest_fiscal_year[0]
                ):
                    latest_fiscal_year = (start_year, value)
                    latest_column = column

            if (
                latest_column is None
                or latest_fiscal_year is None
            ):
                raise ValueError(
                    "No fiscal-year emission-factor column "
                    "was found in the CEA workbook."
                )

            factor_value = sheet.cell(
                row=cls.GRID_FACTOR_ROW,
                column=latest_column,
            ).value

            if factor_value is None:
                raise ValueError(
                    "No grid emission factor was found for "
                    f"fiscal year {latest_fiscal_year[1]}."
                )

            try:
                factor = Decimal(str(factor_value))
            except (InvalidOperation, ValueError) as exc:
                raise ValueError(
                    "CEA emission factor is not a valid decimal value."
                ) from exc

            if factor <= Decimal("0"):
                raise ValueError(
                    "CEA emission factor must be greater than zero."
                )

            effective_from = date(
                latest_fiscal_year[0],
                4,
                1,
            )

            return ImportedEmissionFactor(
                category_name="Electricity",
                factor=factor,
                unit="kgCO2/kWh",
                source=cls.SOURCE_NAME,
                source_version=f"Version {version}",
                effective_from=effective_from,
                effective_to=None,
            )

        finally:
            workbook.close()

    @classmethod
    def get_factors_from_workbook(
        cls,
        workbook_path: str | Path,
    ) -> list[ImportedEmissionFactor]:
        """
        Return all emission factors provided by the CEA workbook.

        The current CEA source adapter provides one normalized
        electricity factor.
        """

        return [
            cls.parse_workbook(workbook_path)
        ]